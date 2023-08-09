#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@Author:Alan
@Create 2023/8/8 21:06
@Version 1.0
"""
import os
import re

import openpyxl

# tablename 字典
table_name_dict = {
    1: "gh_fixed_wages",
    2: "gh_fixed_welfare",
    3: "gh_five_risks",
    4: "gh_accumulation_fund",
    5: "gh_annuity",
    6: "gh_zhe_jiu",
    7: "gh_fixed_car_insurance",
    8: "gh_fix_cailiao",
    9: "gh_traffic_expense",
    10: "gh_item_laowu",
    11: "gh_item_gongnong",
    12: "gh_fixed_water",
    13: "gh_repair",
    14: "gh_item_waigu",
    15: "gh_fixed_taxes",
    16: "",
    17: "gh_fixed_travel_expense",
    18: "gh_fixed_office_allowance",
    19: "gh_fixed_print",
    20: "gh_computer",
    21: "gh_decorate",
    22: "gh_fixed_communication_expense",
    23: "gh_fixed_training_expense",
    24: "gh_fixed_test_fee",
    25: "gh_fixed_labour_protection",
    26: "gh_fixed_agency_fee",
    27: "gh_item_other_expenses",
    28: "gh_fix_trade"
}

# type_classe_dict 字典
file_name_dist = {
    'gongzi': 1,
    'fuli': 2,
    'wuxian': 3,
    'zhufanggongjijin': 4,
    'nianjin': 5,
    'zhejiutanxiao': 6,
    'cheliangbaoxian': 7,
    'cailiaofei': 8,
    'jixiefei': 9,
    'waifulaowu': 10,
    'laowufei': 11,
    'shuidianfei': 12,
    'xiulifei': 13,
    'linshigongzi': 14,
    'shuijin': 15,
    'dizhiyihaopin': 16,
    'chalvfei': 17,
    'yewuzhaodaifei': 18,
    'yinshuafei': 19,
    'weijihaocai': 20,
    'zhuangshifei': 21,
    'tongxunfei': 22,
    'peixunfei': 23,
    'jiancefei': 24,
    'laodongbaohufei': 25,
    'zhongjiefei': 26,
    'qitazhichu': 27,
    'shangmaochengben': 28
}

# 指定要处理的文件夹路径
# folder_path = './testExcelFile'
folder_path = 'D://Projects//gh//slbfNew//28表更改信息//胜利北方公司修改后//建筑公司//测试vba'

# 创建一个新的工作簿用于存储提取的数据
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
# 给汇总表添加表头
new_sheet.append(["id", "type_classe", "tablename"])

print(os.listdir(folder_path))

# 遍历指定文件夹下的所有xlsx文件
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        source_filepath = os.path.join(folder_path, filename)

        # 截取_和.之间的文件名
        pattern = r'(?<=_).*?(?=\.)'
        result = re.search(pattern, filename)

        type_classe_name = result.group()
        # 获取根据文件名获取类型
        type_classe_value = file_name_dist[type_classe_name]

        # 打开源工作簿
        source_workbook = openpyxl.load_workbook(source_filepath)
        source_sheet = source_workbook.active

        # 获取工作簿最大行数
        max_row = source_sheet.max_row

        # 读取第一行，确定id和type_classe所在的列
        header_row = list(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        id_col = 1  # 索引从 1 开始
        type_classe_col = None
        for col_index, header in enumerate(header_row):
            if header == "type_classe":
                type_classe_col = col_index + 1  # 列索引从1开始
        # 遍历除第一行外每一行的第一个单元格确定背景填充色
        for i in range(2, max_row):
            # 获取表格填充色颜色
            color = source_sheet.cell(i, 1).fill.fgColor.rgb
            # 获取黄色行的id和type_classe
            if color == "FFFFFF00":
                # id 截取第一个字符之后的字符串
                id_value = source_sheet.cell(i, id_col).value[1:]
                new_sheet.append([id_value, type_classe_value, table_name_dict[type_classe_value]])

        source_workbook.close()

# 保存新的工作簿到新文件并命名
new_workbook.save('filtered_data.xlsx')
